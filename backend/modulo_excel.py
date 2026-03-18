from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from io import BytesIO

MESI_IT = ["","Gennaio","Febbraio","Marzo","Aprile","Maggio","Giugno",
           "Luglio","Agosto","Settembre","Ottobre","Novembre","Dicembre"]

def genera_modulo_excel(incassi: list, mese: int, anno: int) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Corrispettivi"

    bold = Font(name="Arial", bold=True, size=11)
    normal = Font(name="Arial", size=11)
    bold_gold = Font(name="Arial", bold=True, size=11, color="B5892A")
    thin = Side(style="thin", color="AAAAAA")
    border_bottom = Border(bottom=thin)
    border_full = Border(top=thin, bottom=thin, left=thin, right=thin)
    center = Alignment(horizontal="center")
    right = Alignment(horizontal="right")

    def w(row, col, value, font=None, align=None, border=None, fill=None, num_fmt=None):
        cell = ws.cell(row=row, column=col, value=value)
        if font: cell.font = font
        if align: cell.alignment = align
        if border: cell.border = border
        if fill: cell.fill = fill
        if num_fmt: cell.number_format = num_fmt
        return cell

    w(2, 2, "CITTÀ DI SUZZARA", bold, center)
    w(3, 2, "Prov. Di Mantova", normal, center)
    w(6, 1, "GESTIONE ATTIVITÀ MUSEALE", bold)
    w(7, 1, "INCASSI GIORNALIERI", bold)
    w(9, 1, "mese di", normal)
    w(9, 2, f"{MESI_IT[mese]} {anno}", bold_gold)

    header_font = Font(name="Arial", bold=True, size=10)
    header_fill = PatternFill("solid", fgColor="F0E4C4")
    for col, testo in enumerate(["data","ricevute da n.","ricevute a n.","importo","note – contante/POS"], 1):
        w(11, col, testo, header_font, center, border_full, header_fill)

    row = 12
    totale = 0.0
    for inc in sorted(incassi, key=lambda x: x["data"]):
        from datetime import datetime
        data_str = datetime.fromisoformat(inc["data"]).strftime("%d/%m/%Y")
        w(row, 1, data_str, normal)
        if inc.get("ricevuta_da"): w(row, 2, inc["ricevuta_da"], normal, center)
        if inc.get("ricevuta_a"): w(row, 3, inc["ricevuta_a"], normal, center)
        w(row, 4, inc["importo_totale"], normal, right, num_fmt='#,##0.00')
        nota = inc["modalita"] + (f" — {inc['note']}" if inc.get("note") else "")
        w(row, 5, nota, normal)
        ws.cell(row=row, column=1).border = border_bottom
        ws.cell(row=row, column=4).border = border_bottom
        totale += inc["importo_totale"]
        row += 1

    while row < 42:
        ws.cell(row=row, column=1).border = border_bottom
        row += 1

    w(42, 2, "totale incassato", bold, center)
    tot_cell = w(42, 4, totale, bold, right, border_full, num_fmt='#,##0.00')
    tot_cell.font = Font(name="Arial", bold=True, size=12, color="B5892A")

    w(46, 1, "Suzzara, ___________________")
    w(46, 5, "FIRMA", bold, center)

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 26

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
